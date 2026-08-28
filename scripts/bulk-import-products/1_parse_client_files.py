"""
Step 1 of 3 — run on the HOST (not in odoo shell; the Odoo container has no
xlrd/openpyxl). Parses the client's product-code reference files into a clean
intermediate JSON that step 2 consumes.

Usage:
    python3 1_parse_client_files.py <master.xls> <attributes.xlsx> <out.json>

Expects the same 2-file shape EMG-O has sent twice now (2026-08-25 round):
  - master file: 1 row per product (BL/SL/FG), columns include at minimum
    External ID/Internal Reference (col D, 0-indexed 3), Name (col B, idx 1),
    Product Category (col J, idx 9). Internal Reference pattern must be
    "1-{material-group}-{seq}-{role}" for stone rows (role = BL/SL/FG),
    anything else (e.g. IMP-*/SITE-*) is treated as a non-stone auxiliary
    product and passed through separately.
  - attributes file: same rows, but Slab/FG rows appear twice — once per
    Attribute (col N, idx 13) with its Values (col O, idx 14). We only use
    the "Surface"-named attribute row (Finish); any "Thickness"-named row is
    read but discarded (this system tracks Thickness at the Bundle level,
    not as a product Variant — see the Product/Variant ADR in progress.md).

If a future file uses a different Finish-value vocabulary, extend FINISH_MAP
below — anything not in the map passes through as-is (and step 2 will just
skip it silently if it doesn't match a real product.attribute.value name, so
check step 2's printed skip list after a dry run).
"""
import json
import re
import sys
from collections import Counter

import openpyxl
import xlrd

FINISH_MAP = {
    'Polished / ขัดมัน': 'Polished',
    'Honed / ขัดเรียบ': 'Honed',
    'พ่นทราย / Sand Blasted': 'Sandblasted',
    'พ่นน้ำ / Water Blasted': 'Water Blasted',
    'ปั่นแปรง / Brushed': 'Brushed',
    'พ่นไฟ / Flame Blasted': 'Flame Blasted',
    'พ่นไฟ ปั่นแปรง / Flame and Brushed': 'Flame and Brushed',
}

CATEGORY_KEYWORDS = [
    ('QUARTZITE', 'quartzite'),
    ('LIMESTONE', 'limestone'),
    ('TRAVERTINE', 'limestone'),
    ('ONYX', 'other'),
]


def norm_name(text):
    text = re.sub(r'\s*-\s*(BLOCK|SLAB)\s*$', '', text, flags=re.I)
    text = re.sub(r'\s+', ' ', text).strip()
    return text.replace(' /', '/').replace('/ ', '/')


def strip_code_prefix(full, code):
    full = full.strip()
    if full.upper().startswith(code.upper()):
        full = full[len(code):].strip()
    return full


def main(master_path, attrs_path, out_path):
    wb = xlrd.open_workbook(master_path)
    ws = wb.sheet_by_name(wb.sheet_names()[0])
    rows = [ws.row_values(r) for r in range(1, ws.nrows)]

    groups = {}
    other_rows = []
    for r in rows:
        code = (r[3] or '').strip()
        if not code:
            continue
        norm = code.rstrip('.')
        parts = norm.split('-')
        if len(parts) == 4 and parts[0] == '1' and parts[1].isdigit():
            role = parts[3]
            prefix = '-'.join(parts[:3])
            groups.setdefault(prefix, {})[role] = {'code': norm, 'name': r[1], 'category': r[9]}
        else:
            other_rows.append({'code': norm, 'name': r[1], 'category': r[9], 'uom_file': r[12]})

    materials = []
    skipped = []
    name_conflicts = []
    for prefix, roles in sorted(groups.items()):
        if not all(k in roles for k in ('BL', 'SL', 'FG')):
            skipped.append({'prefix': prefix, 'have': list(roles.keys())})
            continue
        cleaned = {}
        for role in ('BL', 'SL', 'FG'):
            raw = strip_code_prefix(roles[role]['name'], roles[role]['code'])
            raw = re.sub(r'^1-\d-\d+-(BL|SL|FG)\.?\s*', '', raw)  # stray wrong-role prefix left in text
            cleaned[role] = norm_name(raw)

        counts = Counter(cleaned.values())
        canonical, freq = counts.most_common(1)[0]
        if freq < 3:
            name_conflicts.append({'prefix': prefix, 'names': cleaned, 'chosen': canonical})

        cat_text = roles['FG']['category']
        category = 'marble' if cat_text.startswith('Marble') else ('granite' if cat_text.startswith('Granite') else 'other')
        name_upper = canonical.upper()
        for kw, cat in CATEGORY_KEYWORDS:
            if kw in name_upper:
                category = cat
                break

        materials.append({
            'prefix': prefix,
            'name': canonical,
            'category': category,
            'block_code': roles['BL']['code'],
            'slab_code': roles['SL']['code'],
            'fg_code': roles['FG']['code'],
        })

    awb = openpyxl.load_workbook(attrs_path, data_only=True)
    aws = awb[awb.sheetnames[0]]
    finish_by_code = {}
    for row in aws.iter_rows(min_row=2, values_only=True):
        code, attr, vals = row[3], row[13], row[14]
        if code and attr and 'Surface' in attr:
            names = [FINISH_MAP.get(v.strip(), v.strip()) for v in (vals or '').split(',')]
            finish_by_code[code.rstrip('.')] = names

    for m in materials:
        m['finish_values'] = finish_by_code.get(m['slab_code'], [])

    out = {'materials': materials, 'skipped': skipped, 'name_conflicts': name_conflicts, 'other_rows': other_rows}
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print('materials parsed:', len(materials))
    print('skipped (incomplete BL/SL/FG triplet — needs client follow-up):', skipped)
    print('name conflicts (needs a human decision, majority-vote applied as a guess):')
    for c in name_conflicts:
        print(' ', c['prefix'], c['names'], '-> chosen:', c['chosen'])
    print('other (non-stone, e.g. IMP-*/SITE-*) rows:', len(other_rows))
    print('category distribution:', Counter(m['category'] for m in materials))
    print()
    print('Wrote', out_path)


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
