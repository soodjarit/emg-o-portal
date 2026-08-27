import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from qa_data_day_in_life import CATEGORIES

TAG_RE = re.compile(r'<br\s*/?>')
STRIP_RE = re.compile(r'<[^>]+>')

def plain(s):
    s = TAG_RE.sub('\n', s)
    s = STRIP_RE.sub('', s)
    s = s.replace('&amp;', '&').replace('&gt;', '>').replace('&lt;', '<').replace('&quot;', '"')
    return s

def steps_plain(steps):
    return '\n'.join('%d. %s' % (i + 1, plain(st)) for i, st in enumerate(steps))

CHARCOAL = '2B2E31'
BG = 'F3F1EE'
AMBER = 'B8892E'

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor=CHARCOAL)
TITLE_FONT = Font(name='Cambria', bold=True, size=14, color=CHARCOAL)
TITLE_FILL = PatternFill('solid', fgColor=BG)
BODY_FONT = Font(name='Calibri', size=11)
BODY_ALIGN = Alignment(vertical='top', wrap_text=True)
NOTE_FONT = Font(name='Calibri', size=9.5, italic=True, color=AMBER)

COLUMNS = [
    ('TC ID', 12), ('Scenario', 32), ('Precondition', 24), ('Steps', 40),
    ('ข้อมูลตัวอย่าง (Sample Data)', 34), ('Expected Result', 38), ('Priority', 10),
    ('Note', 26), ('Actual Result', 28), ('Status (Pass/Fail)', 16),
    ('Tester', 13), ('Date', 12),
]

wb = openpyxl.Workbook()
ws0 = wb.active
ws0.title = 'Overview'
ws0.sheet_view.showGridLines = False
ws0.column_dimensions['B'].width = 4
ws0.column_dimensions['C'].width = 60
ws0.column_dimensions['D'].width = 40
ws0['C2'] = 'EMG-O — Test Cases: Day-in-the-Life (Order-to-Cash)'
ws0['C2'].font = TITLE_FONT
ws0['C3'] = 'Empire Stone × Empire Granite — stone_slab_inventory + boq_estimation (odoo19-ent-mbx, EE)'
ws0['C3'].font = Font(name='Calibri', italic=True, color='6B6E70')

total = sum(len(c['cases']) for c in CATEGORIES)
prios = {'High': 0, 'Medium': 0, 'Low': 0}
for c in CATEGORIES:
    for tc in c['cases']:
        prios[tc['prio']] += 1

row = 5
ws0.cell(row=row, column=3, value='Total test cases: %d' % total).font = BODY_FONT
row += 1
ws0.cell(row=row, column=3, value='High / Medium / Low: %d / %d / %d' % (prios['High'], prios['Medium'], prios['Low'])).font = BODY_FONT
row += 2
ws0.cell(row=row, column=3, value='หมวด').font = Font(bold=True)
ws0.cell(row=row, column=4, value='จำนวนเคส').font = Font(bold=True)
row += 1
for c in CATEGORIES:
    ws0.cell(row=row, column=3, value='%s — %s' % (c['cat_id'], plain(c['title']).split('. ', 1)[1])).font = BODY_FONT
    ws0.cell(row=row, column=4, value=len(c['cases'])).font = BODY_FONT
    row += 1

for c in CATEGORIES:
    ws = wb.create_sheet(c['cat_id'])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'
    for idx, (name, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value='%s — %s' % (plain(c['title']), plain(c['subtitle']).split(' · ')[0]))
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    ws.row_dimensions[1].height = 26

    for idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    r = 4
    for tc in c['cases']:
        values = [
            tc['id'],
            plain(tc['scenario']),
            plain(tc['pre']),
            steps_plain(tc['steps']),
            plain(tc['sample']),
            plain(tc['expected']),
            tc['prio'],
            plain(tc.get('note', '') or ''),
            '', '', '', '',
        ]
        for idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=idx, value=val if val != '' else None)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
        ws.row_dimensions[r].height = 60
        r += 1

if __name__ == '__main__':
    out = 'emg-o-test-cases-day-in-life.xlsx'
    wb.save(out)
    print('wrote', out)
